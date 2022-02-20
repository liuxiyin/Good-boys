import base64
import openpyxl as op
import pandas
from playwright.sync_api import sync_playwright
from loguru import logger
from email.mime.application import MIMEApplication
from Crypto.Hash import MD5
import requests
import email.mime.multipart
import email.mime.text
import smtplib
import os
import uuid
import pandas as pd
import time
import arrow
import re
import json
import shutil


class exceldeal(object):
    def __init__(self, file):
        self.path_1 = data_file  # 上传的 线下表
        self.path_3 = file  # 下载表
        self.path_4 = os.getcwd() + '\\data\\' + '原数据.xlsx'  # 留存备份，用来对比
        self.path_5 = os.getcwd() + '\\data\\' + '比较.xlsx'  # 比较的结果表，用来邮件发送
        self.path_model = os.getcwd() + '\\data\\' + '导入模板.xlsx'  # 构造一个模板，
        df1 = pandas.read_excel(file)  # 构造 导入模板
        df2 = df1[:][:1]
        df2.to_excel(self.path_model, index=None, sheet_name='薪酬数据')
        self.id_list_updata = []  # 模板有，下载文件（系统）没有
        self.id_list_fill = []  # 下载文件（系统）有，模板没有
        self.id_list_system = []  # 下载文件（系统）比线下多的
        self.flag = 0  # 判断 表头 是不是 两行 ，“0”是，“1”不是

    def excel_rewrite(self):
        path_use1 = os.getcwd() + '\\data\\' + '线下表.xlsx'
        df1 = pd.read_excel(self.path_1, header=1, names=None, dtype=str)
        list1 = df1['姓名'].values.tolist()
        if str(list1[0]) == 'nan':
            self.flag = 1  # 表头 是 两行
        for index, j in df1.iterrows():
            if index < 1:
                continue
            if type(j['姓名']) == str and j['姓名'] != '姓名' and j['姓名'].isdigit()==False:
                pass
            else:
                df1.drop(index, axis=0, inplace=True)
        df1.to_excel(path_use1, index=False)
        wb = op.load_workbook(self.path_1)
        ws = wb.active
        t = ws['A1'].value
        wb.close()
        wb = op.load_workbook(path_use1)
        ws = wb.active
        ws.insert_rows(1)
        ws["A1"] = t
        wb.save(path_use1)
        self.path_1 = path_use1
        wb.close()

    def excel_one_line_to_list(self, uc, num, header=1):  # 线下表导入[模板]用
        df = pd.read_excel(self.path_1, header=header, names=None, dtype=str)
        df.columns = [str(x).replace(" ", "") for x in df.columns]
        uc = uc.replace(" ", "")
        list1 = df[uc].values.tolist()
        list1 = ['nan' if str(x).isspace() else x for x in list1]
        list1 = [str(x).replace(" ", "") for x in list1]
        if header == 1 and self.flag == 1 and str(list1[0]) == 'nan':
            list1.pop(0)
        list1 = [0 if str(x) == 'nan' else x for x in list1]
        wb = op.load_workbook(self.path_model)
        sh = wb["薪酬数据"]
        j = 3
        logger.info(f'字段对应:[{uc}]对应导入模板的[第{num}列]---[{sh.cell(2, num).value}]')
        for i in list1:
            sh.cell(j, num, i)
            j = j + 1
        wb.save(self.path_model)

    def quantity(self, name1, name2):
        """
        1，把模板有，下载文件没有，以及身份证为空的，每行信息取出 【self.id_list_0】
        2，把 下载文件有，模板没有的 每行信息中的【姓名，身份证】填充进去 【self.id_list_1】
        3，以身份证号为参照，填充工号
        """
        df1 = pd.read_excel(name1, header=1, names=None)  # 下载文件
        list11 = df1['工号(必填)'].values.tolist()
        list12 = df1['姓名'].values.tolist()
        list13 = df1['身份证号码'].values.tolist()
        list13 = [str(x) for x in list13]
        print('list13',len(list13), list13)
        df2 = pd.read_excel(name2, header=1, names=None)  # 模板
        delete_num = []
        for index, j in df2.iterrows():  # 检验那一行 没有 身份证号 ，或身份证号不在 系统表 里
            if str(j['身份证号码']) == 'nan' or str(j['身份证号码']) not in list13:
                a = [j['姓名'], str(j['身份证号码'])]
                self.id_list_updata.append(a)
                delete_num.append(index + 3)
                df2.drop(index, axis=0, inplace=True)
        list23 = df2['身份证号码'].values.tolist()
        list23 = [str(x) for x in list23]
        print('list23',len(list23), list23)
        self.id_list_fill = [str(x) for x in list13 if x not in list23]  # 下载文件（系统）有，模板没有
        print('self.id_list_fill', len(self.id_list_fill), self.id_list_fill)
        if len(self.id_list_fill) > 0:  # 如果 有需要填充到 模板 的 个人信息，
            wb = op.load_workbook(name2)  # 打开 [模板]
            ws = wb['薪酬数据']
            t = 0
            for i in delete_num:  # 删除检测到的数据
                i = i - t
                ws.delete_rows(i)
                t = t + 1  # 删除检测到的数据
            row = ws.max_row  # 取得行数
            col = ws.max_column + 1
            list_none = []
            for k, l in zip(list12, list13):
                if str(l) in self.id_list_fill:
                    row = row + 1  # 取得行数加1，最大行后面的一行空行
                    ws.cell(row, 2).value = k
                    ws.cell(row, 3).value = str(l)
                    list_none.append([k, str(l)])
                    for j in range(6, col):
                        ws.cell(row, j).value = 0
            wb.save(name2)
            wb.close()
            logger.info(f'补充的[线下表]缺少的个人信息[姓名,身份证号]({len(list_none)}个):{list_none}')
            self.id_list_system = list_none
        df2 = pd.read_excel(name2, header=1, names=None, dtype=str)  # 模板
        # print(df2.columns)
        list23 = df2['身份证号码'].values.tolist()
        wb = op.load_workbook(name2)  # 打开模板，以身份证为参照，填充工号
        ws = wb['薪酬数据']
        row = 3  # 取得行数
        for i in list23:   # 工号    身份证号
            for k, l in zip(list11, list13):
                if str(l) == str(i):
                    ws.cell(row, 1).value = k
                    row = row+1
                    break
        wb.save(name2)
        wb.close()

    def sort_by_id(self, data1, data2):  # 下载， 模板
        df1 = pd.read_excel(data1, header=1, names=None, dtype=str)  # 下载
        df1 = df1.sort_values('身份证号码')
        df1.to_excel(self.path_4, index=False, sheet_name='薪酬数据')  # 原数据
        df2 = pd.read_excel(data2, header=1, names=None, dtype=str)  # 模板
        df2 = df2.sort_values('身份证号码')
        df2.to_excel(self.path_5, index=False, sheet_name='薪酬数据')  # 比较

    def write_data(self, name1, name2, list_new):  # 读取 按 "身份证" 排序 的 下载文件(系统) 写入 比较
        df1 = pd.read_excel(name1, header=1, dtype=str)
        df1 = df1.sort_values('身份证号码')
        wb = op.load_workbook(name2)
        sh = wb["薪酬数据"]
        for t in list_new:
            # print(t)
            list_1 = df1[t[0]].values.tolist()
            j = 2
            for i in list_1:
                sh.cell(j, t[1], i)
                j = j + 1
        wb.save(name2)

    def compare_data(self, col):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        wb1 = load_workbook(self.path_4)
        ws1 = wb1['薪酬数据']
        wb2 = load_workbook(self.path_5)
        ws2 = wb2['薪酬数据']
        list_row = []
        list_row_yellow = []
        for i, j in zip(ws1[col], ws2[col]):
            if (i.value == 0 and j.value != 0) or (i.value != 0 and j.value == 0):
                list_row_yellow.append(col + str(j.row))  # 黄色
            elif i.value != j.value:
                list_row.append(col + str(j.row))  # 红色
        for i in list_row_yellow:
            ws2[i].fill = PatternFill('solid', fgColor='FFFF00')  # 黄色
        for i in list_row:
            ws2[i].fill = PatternFill('solid', fgColor='FF2D2D')  # 红色
        wb2.save(self.path_5)
        wb1.close()
        wb2.close()

    def color_new(self, path5):  # 姓名与身份证，模板没有的变'9393FF'的颜色，身份证为空与
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        wb1 = load_workbook(path5)
        ws1 = wb1['薪酬数据']
        if len(self.id_list_fill) != 0:
            for i in ws1['C']:
                if str(i.value) in self.id_list_fill:
                    r = 'C' + str(i.row)
                    ws1[r].fill = PatternFill('solid', fgColor='9393FF')
        if len(self.id_list_updata) != 0 or len(self.id_list_system) != 0:
            logger.info(f'[系统]缺少的或者[身份证有问题]的个人信息({len(self.id_list_updata)}个):{self.id_list_updata}')
            row = ws1.max_row
            self.id_list_updata.sort()
            self.id_list_system.sort()
            for d in self.id_list_system:
                # logger.info(f'差异的个人信息（系统）{d}')
                row = row + 1
                ws1.cell(row, 1).value = '（系统）差异信息'
                ws1.cell(row, 2).value = d[0]
                ws1.cell(row, 2).fill = PatternFill('solid', fgColor='9393FF')
                ws1.cell(row, 3).value = d[1]
                ws1.cell(row, 3).fill = PatternFill('solid', fgColor='9393FF')
            for i in self.id_list_updata:
                # logger.info(f'差异的个人信息（线下）{i}')
                row = row + 1
                ws1.cell(row, 1).value = '（线下）差异信息'
                ws1.cell(row, 2).value = i[0]
                ws1['B' + str(row)].fill = PatternFill('solid', fgColor='92CDDC')
                ws1.cell(row, 3).value = i[1]
                ws1['C' + str(row)].fill = PatternFill('solid', fgColor='92CDDC')
        wb1.save(path5)

    def file_deal(self, list1, list2, list_compare):
        self.excel_rewrite()
        for i in list1:
            self.excel_one_line_to_list(uc=i[0], num=i[1], header=i[2])  # 线下表 的内容 保存在 模板 里
        self.quantity(name1=self.path_3, name2=self.path_model)  # 取 下载文件 补全 模板 的姓名，身份证号码,工号
        self.sort_by_id(data1=self.path_3, data2=self.path_model)
        # 按 "身份证" 排序  的 下载文件 放入 原数据 里，按 "身份证" 排序 的[模板(补充完毕)] 放入 比较 里
        self.write_data(name1=self.path_3, name2=self.path_5, list_new=list2)  # 读取 按 "身份证" 排序 的 下载文件 写入 比较
        for col in list_compare:
            self.compare_data(col=col)
        self.color_new(path5=self.path_5)

    def tongjia(self):
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本\n工资', 6, 1], ['基本天数', 9, 1], ['总出勤\n天数', 10, 1], ['正常出勤天数', 11, 1],
                 ['就餐天数', 12, 1], ['事假工时', 13, 1], ['病假工时', 14, 1], ['平时加班时', 15, 1], ['双休加班时', 16, 1],
                 ['国定假加班时', 17, 1], ['夜班天数', 18, 1], ['平时加班工资', 23, 1], ['双休工资', 24, 1], ['国定假工资', 25, 1],
                 ['加班2h就餐天数', 28, 1], ['补助饭贴', 29, 1], ['绩效奖', 30, 1], ['夜班补贴', 33, 1], ['岗位津贴', 34, 1], ['全勤奖', 35, 1],
                 ['中秋福利', 36, 1], ['通讯费', 38, 1], ['高温费', 39, 1], ['公积金补助', 40, 1], ['调整', 41, 1], ['个税专项扣除', 45, 1],
                 ['个人社保', 58, 1], ['外包公司社保', 59, 1], ['个人公积金', 60, 1], ['外包公司公积金', 61, 1], ['外包社保服务费', 73, 1],
                 ['个调税', 81, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['临时工小时计薪标准', 7], ['加班计算基数', 8], ['临时工超10H', 19],
                 ['当月工资', 20], ['加班费（计算）', 21], ['加班费（调整）', 22], ['饭贴标准（月）', 26], ['饭贴标准（天）', 27], ['夜班补贴标准（日）', 31],
                 ['夜班补贴（固定）', 32], ['安全奖', 37], ['其他收入合计', 42], ['补发工资', 43], ['应发工资', 44], ['社保基数', 46], ['公积金基数', 47],
                 ['养老金（个人）', 48], ['养老金（公司）', 49], ['医疗金（个人）', 50], ['医疗金（公司）', 51], ['失业金（个人）', 52], ['失业金（公司）', 53],
                 ['大病医疗（个人）', 54], ['大病医疗（公司）', 55], ['生育保险', 56], ['工伤保险', 57], ['地方附加医疗（公司）', 62], ['社保补缴项（个人）', 63],
                 ['社保补缴项（公司）', 64], ['公积金补缴项（个人）', 65], ['公积金补缴项（公司）', 66], ['个人缴纳（派遣）', 67], ['个人社保手动调整（当月）', 68],
                 ['公司社保手动调整（当月）', 69], ['个人缴纳合计', 70], ['公司缴纳（派遣）', 71], ['公司缴纳合计', 72], ['本期收入手动调整（累计）', 74],
                 ['社保公积金手动调整（累计）', 75], ['专项扣除手动调整（累计）', 76], ['个税手动调整（累计）', 77], ['至本月累计减除手动调整（累计）', 78], ['当月其他扣除', 79],
                 ['个人所得税(自动计算)', 80], ['个人所得税（实际）', 82], ['个税基数', 83], ['税后扣款', 84], ['实发工资', 85], ['人事总成本', 86],
                 ['本期收入（参考）', 87], ['社保公积金（参考）', 88], ['其他减除（参考）', 89], ['专项附加扣除（参考）', 90], ['个税（参考）', 91]]
        list_compare = ['F', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'W', 'X', 'Y', 'AB', 'AC', 'AD', 'AG',
                        'AH', 'AI', 'AJ', 'AL', 'AM', 'AN', 'AO', 'AS', 'BF', 'BG', 'BH', 'BI', 'BU', 'CC']
        self.file_deal(list1, list2, list_compare)

    def shenya(self):
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本工资', 12, 1], ['餐费', 14, 2], ['补贴', 17, 2], ['中夜班补贴', 18, 2],
                 ['其他', 20, 2], ['高温费', 21, 2],  ['公休及病假工资', 22, 2], ['其他扣款', 27, 2], ['病事假', 28, 2],
                 ['病事假', 29, 2], ['个税专项扣除', 31, 1],  ['个人社保', 45, 1], ['公司社保', 46, 1], ['个人公积金', 47, 1],
                 ['公司公积金', 48, 1], ['前锦网络公司社保', 57, 1],  ['社保服务费', 59, 1], ['个调税', 68, 1], ['税后扣款', 71, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['试用期基本工资', 6],
                 ['试用期岗位工资', 7], ['试用期绩效工资', 8], ['基本工资', 9], ['岗位工资', 10], ['绩效工资', 11], ['绩效奖金', 13],
                 ['通讯费', 15], ['交通费', 16], ['加班费（调整）', 19], ['补贴（备增）', 23], ['其他收入合计', 24], ['补发工资', 25],
                 ['补贴（备减）', 26], ['应发工资', 30], ['社保基数', 32], ['公积金基数', 33], ['养老金（个人）', 34], ['养老金（公司）', 35],
                 ['医疗金（个人）', 36], ['医疗金（公司）', 37], ['地方附加医疗（公司）', 38], ['失业金（个人）', 39], ['失业金（公司）', 40],
                 ['大病医疗（个人）', 41], ['大病医疗（公司）', 42], ['生育保险', 43], ['工伤保险', 44], ['社保补缴项（个人）', 49],
                 ['社保补缴项（公司）', 50], ['公积金补缴项（个人）', 51], ['公积金补缴项（公司）', 52], ['个人缴纳（派遣）', 53],
                 ['个人社保手动调整（当月）', 54], ['公司社保手动调整（当月）', 55], ['个人缴纳合计', 56], ['公司缴纳合计', 58], ['固定通讯费', 60],
                 ['本期收入手动调整（累计）', 61], ['社保公积金手动调整（累计）', 62], ['专项扣除手动调整（累计）', 63], ['个税手动调整（累计）', 64],
                 ['至本月累计减除手动调整（累计）', 65], ['当月其他扣除', 66], ['个人所得税(自动计算)', 67], ['个人所得税（实际）', 69], ['个税基数', 70],
                 ['实发工资', 72], ['人事总成本', 73], ['本期收入（参考）', 74], ['社保公积金（参考）', 75], ['其他减除（参考）', 76],
                 ['专项附加扣除（参考）', 77], ['个税（参考）', 78]]
        list_compare = ['L', 'N', 'Q', 'R', 'T', 'U', 'V', 'AA', 'AB', 'AC', 'AE', 'AS', 'AT', 'AU', 'AV',
                        'BE', 'BG', 'BP', 'BS']
        self.file_deal(list1, list2, list_compare)

    def qianjin(self):
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本工资', 6, 1], ['基本天数', 10, 1], ['实际出勤\n天数', 11, 1], ['正常出勤天数', 12, 1],
                  ['病假工时', 13, 1], ['轮休工时', 14, 1], ['就餐天数', 15, 1], ['夜班天数', 16, 1], ['平时\n加班时', 17, 1],
                  ['双休\n加班时', 18, 1], ['国定假加班时', 19, 1], ['平时加班工资', 23, 1], ['双休工资', 24, 1], ['法定假加班工资', 25, 1],
                  ['饭贴', 28, 1], ['绩效奖', 29, 1], ['全勤奖', 31, 1], ['交通补贴', 32, 1], ['通讯费', 33, 1], ['夜班', 34, 1],
                  ['岗位补贴', 35, 1], ['工龄补贴', 36, 1], ['高温补贴', 37, 1], ['专项附加扣除', 43, 1], ['个人社保', 56, 1],
                  ['公司社保', 57, 1], ['个人公积金', 58, 1], ['公司公积金', 59, 1], ['外包社保服务费', 70, 1], ['个调税', 76, 1],
                  ['税后扣款', 81, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['绩效工资', 7], ['拟岗位工资', 8], ['加班计算基数', 9],
                 ['当月工资', 20], ['加班费（计算）', 21], ['加班费（调整）', 22], ['饭贴标准（月）', 26], ['饭贴标准（天）', 27],
                 ['全勤奖标准', 30], ['其他补贴', 38], ['补贴调整', 39], ['其他收入合计', 40], ['补发工资', 41], ['应发工资', 42],
                 ['社保基数', 44], ['公积金基数', 45], ['养老金（个人）', 46], ['养老金（公司）', 47], ['医疗金（个人）', 48],
                 ['医疗金（公司）', 49], ['失业金（个人）', 50], ['失业金（公司）', 51], ['大病医疗（个人）', 52], ['大病医疗（公司）', 53],
                 ['生育保险', 54], ['工伤保险', 55], ['社保补缴项（个人）', 60], ['社保补缴项（公司）', 61], ['公积金补缴项（个人）', 62],
                 ['公积金补缴项（公司）', 63], ['个人缴纳（派遣）', 64], ['个人社保手动调整（当月）', 65], ['公司社保手动调整（当月）', 66],
                 ['个人缴纳合计', 67], ['公司缴纳（派遣）', 68], ['公司缴纳合计', 69], ['本期收入手动调整（累计）', 71], ['社保公积金手动调整（累计）', 72],
                 ['专项扣除手动调整（累计）', 73], ['至本月累计减除手动调整（累计）', 74], ['当月其他扣除', 75], ['个人所得税(自动计算)', 77],
                 ['个税手动调整（累计）', 78], ['个人所得税（实际）', 79], ['个税基数', 80], ['实发工资', 82], ['人事总成本', 83], ['本期收入（参考）', 84],
                 ['社保公积金（参考）', 85], ['其他减除（参考）', 86], ['专项附加扣除（参考）', 87], ['个税（参考）', 88]]
        list_compare = ['F', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'W', 'X', 'Y', 'AB', 'AC',
                        'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AQ', 'BD', 'BE', 'BF', 'BG', 'BR', 'BX', 'CC']
        self.file_deal(list1, list2, list_compare)

    def shenyangku(self):
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本工资', 6, 2], ['岗位工资', 7, 2], ['拟岗位绩效', 8, 1], ['拟岗位', 9, 1],
                 ['迟到', 11, 1], ['应出勤', 12, 1], ['出勤天数', 14, 1], ['调休、年假、公休', 16, 1], ['事假', 17, 1], ['病假', 18, 1],
                 ['缺勤', 19, 1], ['病假', 26, 2], ['其他扣款', 27, 2], ['缺勤扣款', 28, 2], ['事假', 29, 2], ['绩效分', 31, 1],
                 ['绩效', 32, 2], ['餐费', 38, 2], ['通讯费', 39, 2], ['通讯费', 41, 2], ['服务奖', 42, 2], ['技术奖', 43, 2],
                 ['补贴', 45, 2], ['加班费', 48, 2], ['专项附加扣除', 60, 1], ['个人社保', 74, 1], ['公司社保', 75, 1],
                 ['个人公积金', 76, 1], ['外包公司社保', 87, 1], ['社保服务费', 89, 1], ['个调税', 96, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['加班支付标准', 10], ['当月应出勤小时数', 13], ['当月实际出勤小时数', 15],
                 ['当月转正后实际出勤天数', 20], ['当月调薪后实际出勤天数', 21], ['平时加班小时数', 22], ['周末加班小时数', 23], ['法定加班小时数', 24],
                 ['病假折抵工时', 25], ['考勤扣款合计', 30], ['饭贴标准（月）', 33], ['饭贴标准（天）', 34], ['饭贴（计算）', 35], ['加班饭贴', 36],
                 ['饭贴调整', 37], ['房帖', 40], ['油贴（元/月）', 44], ['加班费（计算）', 46], ['加班费（调整）', 47], ['补贴调整', 49],
                 ['转正补差（计薪当月）', 50], ['转正补差调整', 51], ['实际转正补差', 52], ['调薪补差（计薪当月）', 53], ['调薪补差调整', 54],
                 ['实际调薪补差', 55], ['全勤奖', 56], ['其他收入合计', 57], ['补发工资', 58], ['应发工资', 59], ['社保基数', 61], ['公积金基数', 62],
                 ['养老金（个人）', 63], ['养老金（公司）', 64], ['医疗金（个人）', 65], ['医疗金（公司）', 66], ['失业金（个人）', 67], ['失业金（公司）', 68],
                 ['大病医疗（个人）', 69], ['大病医疗（公司）', 70], ['生育保险', 71], ['工伤保险', 72], ['补充工伤', 73], ['公积金（公司）', 77],
                 ['地方附加医疗（公司）', 78], ['社保补缴项（个人）', 79], ['社保补缴项（公司）', 80], ['公积金补缴项（个人）', 81], ['公积金补缴项（公司）', 82],
                 ['个人缴纳（派遣）', 83], ['个人社保手动调整（当月）', 84], ['公司社保手动调整（当月）', 85], ['个人缴纳合计', 86], ['公司缴纳合计', 88],
                 ['社保公积金手动调整（累计）', 90], ['本期收入手动调整（累计）', 91], ['专项扣除手动调整（累计）', 92], ['个税手动调整（累计）', 93],
                 ['至本月累计减除手动调整（累计）', 94], ['当月其他扣除', 95], ['个人所得税(自动计算)', 97], ['个人所得税（实际）', 98], ['个税基数', 99],
                 ['税后扣款', 100], ['实发工资', 101], ['人事总成本', 102], ['本期收入（参考）', 103], ['社保公积金（参考）', 104], ['其他减除（参考）', 105],
                 ['专项附加扣除（参考）', 106], ['个税（参考）', 107]]
        list_compare = ['F', 'G', 'H', 'I', 'K', 'L', 'N', 'P', 'Q', 'R', 'S', 'Z', 'AA', 'AB', 'AC', 'AE', 'AF', 'AL',
                        'AM', 'AO', 'AP', 'AQ', 'AS', 'AV', 'BH', 'BV', 'BW', 'BX', 'CI', 'CK', 'CR']
        self.file_deal(list1, list2, list_compare)

    def shenyang(self):
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本工资', 6, 1], ['总出勤天数', 10, 1], ['总出勤天数', 11, 1], ['正常出勤天数', 12, 1],
                 ['夜班天数', 13, 1], ['平时加班时', 14, 1], ['双休加班时', 15, 1], ['国定假加班时', 16, 1], ['平时加班工资', 20, 1],
                 ['双休工资', 21, 1], ['法定假加班工资', 22, 1], ['饭贴', 25, 1], ['绩效奖', 26, 1], ['交通补贴', 29, 1], ['通讯费', 30, 1],
                 ['夜班', 31, 1], ['岗位补贴', 32, 1], ['其他', 34, 1], ['个税专项扣除', 39, 1], ['个人社保', 52, 1], ['外包公司社保', 53, 1],
                 ['个人公积金', 54, 1], ['外包公司公积金', 55, 1], ['外包社保服务费', 66, 1], ['个调税', 73, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['绩效工资', 7], ['拟岗位工资', 8], ['加班计算基数', 9], ['当月工资', 17],
                 ['加班费（计算）', 18], ['加班费（调整）', 19], ['饭贴标准（月）', 23], ['饭贴标准（天）', 24], ['交通费（元/月）', 27],
                 ['交通费（元/天）', 28], ['高温费', 33], ['补贴调整', 35], ['其他收入合计', 36], ['补发工资', 37], ['应发工资', 38], ['社保基数', 40],
                 ['公积金基数', 41], ['养老金（个人）', 42], ['养老金（公司）', 43], ['医疗金（个人）', 44], ['医疗金（公司）', 45], ['失业金（个人）', 46],
                 ['失业金（公司）', 47], ['大病医疗（个人）', 48], ['大病医疗（公司）', 49], ['生育保险', 50], ['工伤保险', 51], ['社保补缴项（个人）', 56],
                 ['社保补缴项（公司）', 57], ['公积金补缴项（个人）', 58], ['公积金补缴项（公司）', 59], ['个人缴纳（派遣）', 60], ['个人社保手动调整（当月）', 61],
                 ['公司社保手动调整（当月）', 62], ['个人缴纳合计', 63], ['公司缴纳（派遣）', 64], ['公司缴纳合计', 65], ['本期收入手动调整（累计）', 67],
                 ['社保公积金手动调整（累计）', 68], ['专项扣除手动调整（累计）', 69], ['个税手动调整（累计）', 70], ['至本月累计减除手动调整（累计）', 71],
                 ['当月其他扣除', 72], ['个人所得税(自动计算)', 74], ['个人所得税（实际）', 75], ['个税基数', 76], ['税后扣款', 77], ['实发工资', 78],
                 ['人事总成本', 79], ['本期收入（参考）', 80], ['社保公积金（参考）', 81], ['其他减除（参考）', 82], ['专项附加扣除（参考）', 83],
                 ['个税（参考）', 84]]
        list_compare = ['F', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'T', 'U', 'V', 'Y', 'Z', 'AC', 'AD', 'AE', 'AF', 'AH',
                        'AM', 'AZ', 'BA', 'BB', 'BC', 'BN', 'BU']
        self.file_deal(list1, list2, list_compare)

    def jiayang(self):
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本工资', 6, 1], ['总出勤\n天数', 9, 1], ['总出勤\n天数', 11, 1],
                 ['正常出勤天数', 12, 1], ['餐补天数', 13, 1], ['轮休工时', 14, 1], ['病事婚丧产工', 15, 1], ['平时\n加班时', 16, 1],
                 ['双休\n加班时', 17, 1], ['国定加班时', 18, 1], ['夜班天数', 19, 1], ['延迟餐补天数', 20, 1], ['早餐补贴（7:00前）', 21, 1],
                 ['A延班补贴（18:30）', 22, 1], ['B延班补贴  （至凌晨3点）', 23, 1], ['平时加班工资', 28, 1], ['双休加班工资', 29, 1],
                 ['法定加班工资', 30, 1], ['餐费补贴', 34, 1], ['夜班津贴', 37, 1], ['工龄+岗位', 38, 1], ['全勤奖', 40, 1],
                 ['个人考核/绩效奖', 41, 1], ['团队考核', 42, 1], ['安全质量奖', 43, 1], ['住房补贴', 44, 1], ['岗位补贴/技能津贴', 45, 1],
                 ['工作质量奖', 46, 1], ['通讯补贴', 47, 1], ['工会费', 49, 1], ['高温费', 50, 1], ['福利费', 51, 1], ['调整', 52, 1],
                 ['个税专项扣除', 56, 1], ['个人社保', 69, 1], ['外包公司社保', 70, 1], ['个人公积金', 71, 1], ['外包公司公积金', 72, 1],
                 ['劳防用品', 73, 1], ['外包社保服务费', 82, 1], ['个调税', 90, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['临时工小时计薪标准', 7], ['加班计算基数', 8], ['当月应出勤小时数（临时工）', 10],
                 ['当月工资1', 24], ['当月工资', 25], ['加班费（计算）', 26], ['加班费（调整）', 27], ['饭贴标准（月）', 31], ['饭贴标准（天）', 32],
                 ['饭贴调整', 33], ['夜班补贴标准（日）', 35], ['夜班补贴（固定）', 36], ['岗位补贴', 39], ['节日补贴', 48], ['其他收入合计', 53],
                 ['补发工资', 54], ['应发工资', 55], ['社保基数', 57], ['公积金基数', 58], ['养老金（个人）', 59], ['养老金（公司）', 60],
                 ['医疗金（个人）', 61], ['医疗金（公司）', 62], ['失业金（个人）', 63], ['失业金（公司）', 64], ['大病医疗（个人）', 65],
                 ['大病医疗（公司）', 66], ['生育保险', 67], ['工伤保险', 68], ['社保补缴项（个人）', 74], ['社保补缴项（公司）', 75],
                 ['公积金补缴项（个人）', 76], ['公积金补缴项（公司）', 77], ['个人缴纳（派遣）', 78], ['个人缴纳合计', 79], ['公司缴纳（派遣）', 80],
                 ['公司缴纳合计', 81], ['本期收入手动调整（累计）', 83], ['社保公积金手动调整（累计）', 84], ['专项扣除手动调整（累计）', 85],
                 ['个税手动调整（累计）', 86], ['至本月累计减除手动调整（累计）', 87], ['当月其他扣除', 88], ['个人所得税(自动计算)', 89],
                 ['个人所得税（实际）', 91], ['个税基数', 92], ['税后扣款', 93], ['实发工资', 94], ['人事总成本', 95], ['本期收入（参考）', 96],
                 ['社保公积金（参考）', 97], ['其他减除（参考）', 98], ['专项附加扣除（参考）', 99], ['个税（参考）', 100]]
        list_compare = ['F', 'I', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'AB', 'AC', 'AD',
                        'AH', 'AK', 'AL', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AW', 'AX', 'AY', 'AZ', 'BD',
                        'BQ', 'BR', 'BS', 'BT', 'BU', 'CD', 'CL']
        self.file_deal(list1, list2, list_compare)

    def liuzhou(self):
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本工资', 6, 1], ['总出勤\n天数', 9, 1], ['总出勤\n天数', 10, 1], ['正常出勤天数', 12, 1],
                 ['就餐天数', 13, 1], ['平时加班时', 14, 1], ['双休加班时', 15, 1], ['国定假加班时', 16, 1], ['夜班出勤天数', 17, 1],
                 ['平时加班工资', 22, 1], ['双休加班工资', 23, 1], ['法定加班工资', 24, 1], ['节日补贴', 25, 1], ['饭贴', 28, 1], ['绩效', 29, 1],
                 ['交通补贴', 32, 1], ['夜班补贴', 35, 1], ['岗位奖', 36, 1], ['全勤', 37, 1], ['安全', 38, 1], ['调整', 40, 1],
                 ['个税专项扣除', 44, 1], ['个人社保', 57, 1], ['外包公司社保', 58, 1], ['个人公积金', 59, 1], ['外包公司公积金', 60, 1],
                 ['外包社保服务费', 71, 1], ['个调税', 79, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['临时工小时计薪标准', 7], ['加班计算基数', 8], ['当月实际出勤小时数', 11],
                 ['临时工超10H', 18], ['当月工资', 19], ['加班费（计算）', 20], ['加班费（调整）', 21], ['饭贴标准（月）', 26], ['饭贴标准（天）', 27],
                 ['交通费（元/月）', 30], ['交通费（元/天）', 31], ['夜班补贴标准（日）', 33], ['夜班补贴（固定）', 34], ['高温费', 39], ['其他收入合计', 41],
                 ['补发工资', 42], ['应发工资', 43], ['社保基数', 45], ['公积金基数', 46], ['养老金（个人）', 47], ['养老金（公司）', 48],
                 ['医疗金（个人）', 49], ['医疗金（公司）', 50], ['失业金（个人）', 51], ['失业金（公司）', 52], ['大病医疗（个人）', 53],
                 ['大病医疗（公司）', 54], ['生育保险', 55], ['工伤保险', 56], ['社保补缴项（个人）', 61], ['社保补缴项（公司）', 62],
                 ['公积金补缴项（个人）', 63], ['公积金补缴项（公司）', 64], ['个人缴纳（派遣）', 65], ['个人社保手动调整（当月）', 66],
                 ['公司社保手动调整（当月）', 67], ['个人缴纳合计', 68], ['公司缴纳（派遣）', 69], ['公司缴纳合计', 70], ['本期收入手动调整（累计）', 72],
                 ['社保公积金手动调整（累计）', 73], ['专项扣除手动调整（累计）', 74], ['个税手动调整（累计）', 75], ['至本月累计减除手动调整（累计）', 76],
                 ['当月其他扣除', 77], ['个人所得税(自动计算)', 78], ['个人所得税（实际）', 80], ['个税基数', 81], ['税后扣款', 82], ['实发工资', 83],
                 ['人事总成本', 84], ['本期收入（参考）', 85], ['社保公积金（参考）', 86], ['其他减除（参考）', 87], ['专项附加扣除（参考）', 88], ['个税（参考）', 89]]
        list_compare = ['F', 'I', 'J', 'L', 'M', 'N', 'O', 'P', 'Q', 'V', 'W', 'X', 'Y', 'AB', 'AC', 'AF', 'AI', 'AJ',
                        'AK', 'AL', 'AN', 'AR', 'BE', 'BF', 'BG', 'BH', 'BS', 'CA']
        self.file_deal(list1, list2, list_compare)

    def moyulu(self):
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本工资', 6, 1], ['基本天数', 11, 1], ['总出勤\n天数', 12, 1], ['当月实际出勤小时数', 13, 1],
                 ['正常出勤天数', 14, 1], ['事假工时', 15, 1], ['病假工时', 16, 1], ['就餐天数', 19, 1], ['平时加班时', 20, 1],
                 ['双休加班时', 21, 1], ['国定假加班时', 22, 1], ['夜班天数', 23, 1], ['加班2h就餐天数', 24, 1], ['平时加班工资', 29, 1],
                 ['双休工资', 30, 1], ['国定工资', 31, 1], ['补助饭贴', 33, 1], ['绩效', 34, 1], ['夜班补贴', 35, 1], ['岗位奖', 36, 1],
                 ['高温费', 37, 1], ['年终奖', 38, 1], ['调整', 39, 1], ['个税专项扣除', 43, 1], ['个人社保', 56, 1], ['公司社保', 57, 1],
                 ['个人公积金', 58, 1], ['公司公积金', 59, 1], ['外包服务费', 70, 1], ['个调税', 78, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['绩效工资', 7], ['拟岗位工资', 8], ['临时工小时计薪标准', 9], ['加班计算基数', 10],
                 ['轮休工时', 17], ['当月应出勤小时数（临时工）', 18], ['临时工超10H', 25], ['当月工资', 26], ['加班费（计算）', 27], ['加班费（调整）', 28],
                 ['饭贴标准（天）', 32], ['其他收入合计', 40], ['补发工资', 41], ['应发工资', 42], ['社保基数', 44], ['公积金基数', 45], ['养老金（个人）', 46],
                 ['养老金（公司）', 47], ['医疗金（个人）', 48], ['医疗金（公司）', 49], ['失业金（个人）', 50], ['失业金（公司）', 51], ['大病医疗（个人）', 52],
                 ['大病医疗（公司）', 53], ['生育保险', 54], ['工伤保险', 55], ['社保补缴项（个人）', 60], ['社保补缴项（公司）', 61], ['公积金补缴项（个人）', 62],
                 ['公积金补缴项（公司）', 63], ['个人缴纳（派遣）', 64], ['个人社保手动调整（当月）', 65], ['公司社保手动调整（当月）', 66], ['个人缴纳合计', 67],
                 ['公司缴纳（派遣）', 68], ['公司缴纳合计', 69], ['本期收入手动调整（累计）', 71], ['社保公积金手动调整（累计）', 72], ['专项扣除手动调整（累计）', 73],
                 ['个税手动调整（累计）', 74], ['至本月累计减除手动调整（累计）', 75], ['当月其他扣除', 76], ['个人所得税(自动计算)', 77], ['个人所得税（实际）', 79],
                 ['个税基数', 80], ['税后扣款', 81], ['实发工资', 82], ['人事总成本', 83], ['本期收入（参考）', 84], ['社保公积金（参考）', 85],
                 ['其他减除（参考）', 86], ['专项附加扣除（参考）', 87], ['个税（参考）', 88]]
        list_compare = ['F', 'K', 'L', 'M', 'N', 'O', 'P', 'S', 'T', 'U', 'V', 'W', 'X', 'AC', 'AD', 'AE', 'AG', 'AH',
                        'AI', 'AJ', 'AK', 'AL', 'AM', 'AQ', 'BD', 'BE', 'BF', 'BG', 'BR', 'BZ']
        self.file_deal(list1, list2, list_compare)

    def moyulu_qita(self):
        df_1 = pandas.read_excel(self.path_1, header=1)
        a = [0 if str(x) == 'nan' else x for x in df_1['优秀'].values.tolist()]
        b = [0 if str(x) == 'nan' else x for x in df_1['延锋补助激励'].values.tolist()]
        c = [0 if str(x) == 'nan' else x for x in df_1['待岗补助'].values.tolist()]
        d = []
        for i, j, k in zip(a, b, c):
            i = i + j + k
            d.append(i)
        wb = op.load_workbook(self.path_1)
        ws = wb.active
        num = 2
        for i in d:
            num = num + 1
            ws.cell(num, 26).value = i
        wb.save(self.path_1)
        wb.close()
        list1 = [['姓名', 2, 1], ['身份证号', 3, 1], ['基本\n工资', 6, 1], ['基本天数', 11, 1], ['总出勤\n天数', 12, 1], ['正常出勤天数', 14, 1],
                 ['病假工时', 16, 1], ['轮休工时', 17, 1], ['就餐天数', 19, 1], ['平时加班时', 20, 1], ['双休加班时', 21, 1], ['国定假加班时', 22, 1],
                 ['夜班工时（天）', 23, 1], ['加班2h就餐天数', 24, 1], ['白班超过10小时', 25, 1], ['平时加班工资', 29, 1], ['双休工资', 30, 1],
                 ['国定工资', 31, 1], ['饭贴', 33, 1], ['绩效', 34, 1], ['夜班补贴', 35, 1],
                 ['优秀', 38, 1], ['调整', 39, 1],  ['个人社保', 56, 1], ['公司社保', 57, 1], ['外包服务费', 70, 1]]
        list2 = [['计薪统计月', 4], ['发薪日期', 5], ['绩效工资', 7], ['拟岗位工资', 8], ['临时工小时计薪标准', 9],
                 ['加班计算基数', 10], ['当月实际出勤小时数', 13], ['事假工时', 15], ['当月应出勤小时数（临时工）', 18], ['当月工资', 26],
                 ['加班费（计算）', 27], ['加班费（调整）', 28], ['饭贴标准（天）', 32], ['岗位补贴', 36], ['高温费', 37], ['其他收入合计', 40],
                 ['补发工资', 41], ['应发工资', 42], ['专项扣除', 43], ['社保基数', 44], ['公积金基数', 45], ['养老金（个人）', 46],
                 ['养老金（公司）', 47], ['医疗金（个人）', 48], ['医疗金（公司）', 49], ['失业金（个人）', 50], ['失业金（公司）', 51],
                 ['大病医疗（个人）', 52], ['大病医疗（公司）', 53], ['生育保险', 54], ['工伤保险', 55], ['公积金（个人）', 58],
                 ['公积金（公司）', 59], ['社保补缴项（个人）', 60], ['社保补缴项（公司）', 61], ['公积金补缴项（个人）', 62], ['公积金补缴项（公司）', 63],
                 ['个人缴纳（派遣）', 64], ['个人社保手动调整（当月）', 65], ['公司社保手动调整（当月）', 66], ['个人缴纳合计', 67], ['公司缴纳（派遣）', 68],
                 ['公司缴纳合计', 69], ['本期收入手动调整（累计）', 71], ['社保公积金手动调整（累计）', 72], ['专项扣除手动调整（累计）', 73],
                 ['个税手动调整（累计）', 74], ['至本月累计减除手动调整（累计）', 75], ['当月其他扣除', 76], ['个人所得税(自动计算)', 77],
                 ['个税手动调整（当月）', 78], ['个人所得税（实际）', 79], ['个税基数', 80], ['税后扣款', 81], ['实发工资', 82], ['人事总成本', 83],
                 ['本期收入（参考）', 84], ['社保公积金（参考）', 85], ['其他减除（参考）', 86], ['专项附加扣除（参考）', 87], ['个税（参考）', 88]]
        list_compare = ['F', 'K', 'L', 'N', 'P', 'Q', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'AC', 'AD', 'AE', 'AG', 'AH',
                        'AI', 'AL', 'AM', 'BD', 'BE', 'BR']
        self.file_deal(list1, list2, list_compare)

    def main(self, str_name):
        if '统嘉' in str_name:
            self.tongjia()
        elif '申雅' in str_name:
            self.shenya()
        elif '前锦广州' in str_name or '广州龙仁' in str_name:
            self.qianjin()
        elif '沈阳库基层' in str_name:
            self.shenyangku()
        elif '沈阳' in str_name:
            self.shenyang()
        elif '嘉扬' in str_name:
            self.jiayang()
        elif '柳州' in str_name:
            self.liuzhou()
        elif '入场安亭墨玉路' in str_name:
            self.moyulu()
        elif '墨玉路-其他' in str_name:
            self.moyulu_qita()


class ZM(object):
    def __init__(self, playwright):
        self.downloads_path = os.getcwd() + '\\data\\'
        path1 = os.getcwd() + '\\data\\'
        if not os.path.exists(path1):
            os.mkdir(path1)
        else:
            list_dir = os.listdir(path1)
            for i in list_dir:
                file_name = os.path.join(path1 + i)
                if os.path.isfile(file_name):
                    os.remove(file_name)
                elif os.path.isdir(file_name):
                    shutil.rmtree(file_name)
        self.playwright = playwright
        self.browser_type = self.playwright.chromium
        self.browser = self.browser_type.launch(headless=False)
        self.context = self.browser.new_context(accept_downloads=True)
        self.page = self.context.new_page()

    def send_email(self):
        """"
        发送邮件
        """
        recipientAddrs = setting.get("recipientAddrs")
        smtpHost = 'smtp.163.com'
        port = 465
        sendAddr = '18016454917@163.com'
        password = 'EPNZMCUJPPMCCVRJ'
        msg = email.mime.multipart.MIMEMultipart()
        msg['from'] = sendAddr  # 发送人邮箱地址
        msg['to'] = recipientAddrs  # 多个收件人的邮箱应该放在字符串中,用字符分隔, 然后用split()分开,不能放在列表中, 因为要使用encode属性
        msg['subject'] = "比较文件"
        content = "比较文件"  # 内容
        txt = email.mime.text.MIMEText(content, 'plain', 'utf-8')
        msg.attach(txt)
        logger.info('准备添加附件....')
        part = MIMEApplication(open(os.getcwd() + '\\data\\' + '比较.xlsx', 'rb').read())
        file_name = os.path.split(os.getcwd() + '\\data\\' + '比较.xlsx')[-1]
        part.add_header('Content-Disposition', 'attachment', filename=file_name)  # 给附件重命名,一般和原文件名一样,改错了可能无法打开.
        msg.attach(part)
        logger.info("附件添加成功")
        smtp = smtplib.SMTP_SSL(smtpHost, port)  # 需要一个安全的连接，用SSL的方式去登录得用SMTP_SSL，之前用的是SMTP（）.端口号465或587
        smtp.login(sendAddr, password)  # 发送方的邮箱，和授权码（不是邮箱登录密码）
        smtp.sendmail(sendAddr, recipientAddrs.split(";"), str(msg))  # 注意, 这里的收件方可以是多个邮箱,用";"分开, 也可以用其他符号
        smtp.quit()
        logger.info('邮件发送成功')

    def login(self):
        self.page.goto('http://103.156.68.194:8082/zm')
        time.sleep(1)
        self.page.fill('//*[@id="ContentHtml_txtUserName"]', 'RPA')
        self.page.fill('//*[@id="txtPassword"]', '111111')
        self.page.click('//*[@id="ContentHtml_btnLogin"]')
        self.page.wait_for_load_state(state="networkidle", timeout=1000*60*2)
        self.page.hover('//*[@id="slidebar2"]')
        time.sleep(3)
        self.page.click('//*[@id="app-menu-nav-hr"]/li/label/span')
        time.sleep(1)
        self.page.click('//*[@id="app-menu-nav-hr"]/li/ul/li[1]/label/span')
        time.sleep(1)
        self.page.click('//*[@id="app-menu-nav-hr"]/li/ul/li[1]/ul/li[1]/a/span')
        self.page.hover('//*[@id="DivMyCalendar"]')
        logger.info('打开薪酬计算界面')
        time.sleep(2)

    def name_get(self):
        df1 = pd.read_excel(data_file, header=1, names=None)
        if '身份证号' not in df1.columns:
            logger.info('没有发现"身份证号"这一列，请检查一遍')
            logger.info('表头放在[A1],确保有一列为“身份证号”,并且“身份证号”与“姓名”一样放在[第2行],没有"身份证号"的一行不能录入')
            raise
        xl = op.load_workbook(data_file)
        sheet_names = xl.sheetnames
        a1 = xl[sheet_names[0]]["A1"].value
        flag = 0
        if '校验' in a1:
            logger.info('进入校验流程')
            flag = 1
        logger.info(f'{a1}')
        pattern = re.compile(r'\d+')  # 查找数字
        l = pattern.findall(a1)
        if len(l) < 2:
            logger.info('未在[A1]中搜寻到日期')
            raise
        time_year = str(l[0])
        time_month = str(l[1])
        name = 'False'
        dict_ruler = {  # 读取字典，与 日期 组合成 name
            '申雅': '申雅', '沈阳库基层': '沈阳库基层',
            '前锦': '前锦广州', '广州龙仁': '广州龙仁',
            '沈阳基层': '沈阳', '嘉扬': '嘉扬', '柳州': '柳州',
            '入场安亭': '安亭统嘉', '华域': '延锋A组统嘉',
            '入场墨玉路': '入场安亭墨玉路', '入场基层': '墨玉路-其他'}
        for i in dict_ruler:
            if i in a1:
                name = l[0] + '年' + l[1] + '月' + dict_ruler[i]
                break
        if name == 'False':
            logger.info('名称获取失败')
            raise
        logger.info('表格读取成功')
        return name, time_year, time_month, flag

    def select_time(self, main_iframe, time_year, time_month, time_day, xpath):
        if int(time_month) < 1:
            time_year = str(int(time_year) - 1)
            time_month = str(int(time_month) + 12)
        elif int(time_month) > 12:
            time_year = str(int(time_year) + 1)
            time_month = str(int(time_month) - 12)
        month_dict = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10, "十一": 11,
                      "十二": 12}
        time.sleep(1)
        main_iframe.click(xpath)
        for i in range(50):
            time.sleep(0.3)
            t = main_iframe.text_content('//th[@title="选择月份"]')
            if int(t[-4:]) == int(time_year) and month_dict[t[:-6]] == int(time_month):
                break
            if int(t[-4:]) > int(time_year):
                main_iframe.click(xpath + '/div/div/ul/li[1]/div/div[1]/table/thead/tr[1]/th[1]')
            elif int(t[-4:]) < int(time_year):
                main_iframe.click(xpath + '/div/div/ul/li[1]/div/div[1]/table/thead/tr[1]/th[3]')
            elif month_dict[t[:-6]] > int(time_month):
                main_iframe.click(xpath + '/div/div/ul/li[1]/div/div[1]/table/thead/tr[1]/th[1]')
            elif month_dict[t[:-6]] < int(time_month):
                main_iframe.click(xpath + '/div/div/ul/li[1]/div/div[1]/table/thead/tr[1]/th[3]')
        if len(time_month) < 2:
            time_month = '0' + time_month
        if len(time_day) < 2:
            time_day = '0' + time_day
        main_iframe.click(f'//td[@data-day="{time_year}-{time_month}-{time_day}"]')
        time.sleep(1)
        logger.info(f'选择日期为：{time_year}-{time_month}-{time_day}')

    def select_ruler(self, name, main_iframe):
        logger.info('选择计算规则')
        main_iframe.click('//*[@id="button_ctl00_main_frmMain___CalculationRule"]/i')
        frm_iframe = self.page.wait_for_selector('//*[@id="frmList"]').content_frame()
        frm_iframe.select_option('//select[@class="form-control knxDropDownList dropdown-pagesize"]', "200")
        time.sleep(2)
        find_ruler = 'False'
        dict_ruler = {'申雅': '汽配-申雅', '沈阳库基层': '薪资规则-5_15',
                      '前锦广州': '汽配-前锦广州', '广州龙仁': '汽配-前锦广州',
                      '沈阳': '汽配-沈阳', '嘉扬': '汽配-嘉扬', '柳州': '汽配-柳州',
                      '安亭统嘉': '汽配-统嘉', '延锋A组统嘉': '汽配-统嘉',
                      '入场安亭墨玉路': '汽配-墨玉路、其他', '墨玉路-其他': '汽配-墨玉路、其他'}
        for i in dict_ruler:  # 读取字典，获取要选择的计算规则
            if i in name:
                find_ruler = dict_ruler[i]
                break
        if find_ruler == 'False':
            logger.info('未匹配到要求的计算规则')
            raise
        logger.info(f'选择的计算规则是：{find_ruler}')
        frm_iframe.click(f'//div[text()="{find_ruler}"]/../../td[6]/div/a')
        time.sleep(2)

    def select_employee(self, name, main_iframe):
        logger.info('开始选择员工')
        main_iframe.click('//*[@id="button_ctl00_main_frmMain___EmpIds"]/i')
        frm_iframe = self.page.wait_for_selector('//*[@id="frmListCommon"]').content_frame()
        time.sleep(0.5)
        frm_iframe.click('//*[@id="btnFilter1"]')
        while frm_iframe.is_visible('//*[@id="solution_setting"]/fieldset/div/table/tbody/tr[1]/td[7]/a'):
            frm_iframe.click('//*[@id="solution_setting"]/fieldset/div/table/tbody/tr[1]/td[7]/a')
            time.sleep(0.3)
        frm_iframe.click('//*[@id="Button3"]')
        time.sleep(2)  # 接下来选择薪酬计算员工，等于
        frm_iframe.select_option('//*[@id="solution_setting"]/fieldset/div/table/tbody/tr/td[2]/select',
                                 label='薪资规则 (员工基本信息表)')
        frm_iframe.select_option('//*[@id="solution_setting"]/fieldset/div/table/tbody/tr/td[3]/select', label='等于')
        find_ruler = 'False'
        dict_ruler = {'申雅': '汽配-申雅', '沈阳库基层': '薪资规则5-15（汽配-沈阳）',
                      '前锦广州': '汽配-前锦广州、东莞、龙仁', '广州龙仁': '汽配-前锦广州、东莞、龙仁',
                      '沈阳': '汽配-沈阳', '嘉扬': '汽配-嘉扬', '柳州': '汽配-柳州',
                      '安亭统嘉': '汽配-统嘉', '延锋A组统嘉': '汽配-统嘉',
                      '入场安亭墨玉路': '汽配-墨玉路、其他', '墨玉路-其他': '汽配-墨玉路、其他'}  # 读取字典，获取选择的规则
        for i in dict_ruler:
            if i in name:
                find_ruler = dict_ruler[i]
                logger.info(f'员工选择：{find_ruler}')
                break
        if find_ruler == 'False':
            logger.info('未匹配到要求的员工选择计算规则')
            raise
        frm_iframe.select_option('//*[@id="solution_setting"]/fieldset/div/table/tbody/tr/td[4]/div[1]/select',
                                 label=find_ruler)
        time.sleep(2)
        frm_iframe.click('//input[@value="查询"]')
        time.sleep(2)
        frm_iframe.click('//input[@value="全部选择"]')
        time.sleep(2)
        logger.info('选择完毕')

    def new_create(self, name, time_year, time_month, main_iframe):
        logger.info('打开新增界面')
        main_iframe.click('//a[text()="新增"]')
        time.sleep(1)
        main_iframe.fill('//*[@id="ctl00_main_frmMain___CalculateName_Chinese_editor"]', name)  # 填充新增名称
        # 选择计薪统计月
        main_iframe.click('//*[@id="button_ctl00_main_frmMain___PayrollMonth"]/i')
        frm_iframe = self.page.wait_for_selector('//*[@id="frmList"]').content_frame()
        frm_iframe.select_option('//*[@id="main_UpdatePanel1"]/div[2]/ul/li[6]/span/select', "200")
        time.sleep(1)
        if len(time_month) < 2:
            time_month = '0'+time_month
        logger.info(f'计薪统计月为{time_year}年{time_month}月')
        frm_iframe.click(
            '//a[@onclick="ButtonConfim(' + "'" + time_year + time_month + "','" + time_year + "/" + time_month + "')" + '"' + "]")
        # 日期选择 发薪日期、计薪开始日期、计薪结束日期
        logger.info('依次选择 发薪日期、计薪开始日期、计薪结束日期')
        self.select_time(main_iframe, time_year, str(int(time_month) + 1), '15', '//*[@id="formEdit"]/div[3]/div/div')
        self.select_time(main_iframe, time_year, str(int(time_month) - 1), '26', '//*[@id="formEdit"]/div[4]/div/div')
        self.select_time(main_iframe, time_year, time_month, '25', '//*[@id="formEdit"]/div[5]/div/div')

        self.select_ruler(name, main_iframe)  # 计算规则
        self.select_employee(name, main_iframe)  # 员工选择
        time.sleep(2)
        main_iframe.click('//input[@value="保存"]')  # 保存
        logger.info('保存新建项目')
        time.sleep(10)

    def all_compute(self, main_iframe):
        logger.info('点击全部计算')
        main_iframe.click('//input[@value="全部计算"]')
        main_iframe.wait_for_selector('//span[text()="已完成"]', timeout=1000 * 60 * 3)
        self.page.click('//button[text()="确定"]')
        time.sleep(2)
        main_iframe.click('//*[@id="divProcess"]/div/div/div[1]/a/i')

    def file_download(self, name, main_iframe):
        main_iframe.wait_for_selector('//*[@id="form1"]/div[3]/div[1]/span[3]/ul/li[7]/span/select', timeout=1000*60*3)
        main_iframe.select_option('//*[@id="form1"]/div[3]/div[1]/span[3]/ul/li[7]/span/select', "200")
        time.sleep(1)
        main_iframe.click(f'//td[5]/div/span[text()="{name}"]/../div')
        time.sleep(1)
        main_iframe.click(f'//td[5]/div/span[text()="{name}"]/../div/ul/li/a[text()="维护"]')
        time.sleep(3)
        logger.info('进入维护界面')
        self.all_compute(main_iframe)
        time.sleep(2)
        main_iframe.click('//a[text()="导入与导出"]')
        # 文件下载
        logger.info('下载导出文件')
        with self.page.expect_download(timeout=1000 * 60) as download_info:
            main_iframe.click('//input[@value="导出模板含员工数据"]')
        download = download_info.value
        file_name = download.suggested_filename
        download.save_as(self.downloads_path + file_name)
        time.sleep(5)
        return self.downloads_path + file_name

    def update_excel(self, main_iframe, flag):
        if flag == 0:  # 不是校验文件,上传文件
            logger.info('上传文件')
            with self.page.expect_file_chooser() as fc_info:
                main_iframe.click('//i[@class="vx-icon-wenjian"]')
            file_chooser = fc_info.value
            file_chooser.set_files(os.path.join(os.getcwd() + '\\data\\' + '导入模板.xlsx'))
        main_iframe.wait_for_selector('//div[@id="ImportMessage"]/span[@title]', timeout=1000*50)
        if main_iframe.is_visible('//div[@id="ImportMessage"]/span[@title]/../a'):  # 上传失败显示的【详情】
            main_iframe.click('//div[@id="ImportMessage"]/span[@title]/../a')
            text1 = self.page.inner_text('//*[@id="TableDialog"]/div[2]/table/tbody')
            logger.info(text1)
            time.sleep(3)
            if self.page.is_visible('//*[@id="vxDialog"]/div/div/div[1]/a/i'):
                self.page.click('//*[@id="vxDialog"]/div/div/div[1]/a/i')
                time.sleep(2)
        text2 = main_iframe.inner_text('//div[@id="ImportMessage"]/span[@title]')
        logger.info(text2)
        if '错误' in text2:
            raise
        time.sleep(2)
        main_iframe.click('//input[@value="取消"]')
        time.sleep(3)
        self.all_compute(main_iframe)
        time.sleep(2)
        main_iframe.click('//input[@value="保存"]')  # 保存
        logger.info('点击保存')
        self.page.wait_for_selector('//button[text()="确定"]', timeout=1000 * 60)
        self.page.click('//button[text()="确定"]')
        time.sleep(5)
        main_iframe.click('//input[@value="全部计算"]/../input[@value="取消"]')
        logger.info('回归薪酬计算界面')

    def updata_file(self, strs, nass_name='文件'):
        from loguru import logger
        # robotNo = serial_no
        # robotPwd = robot_secret

        obj = MD5.new()
        obj.update(robot_secret.encode("utf-8"))  # gb2312 Or utf-8
        robotPwd = obj.hexdigest()

        url = f'{base_url}/internal/file/upload/result'
        # url = f'{base_url}/file/upload/{robotNo}/{robotPwd}'
        header = {
            'robotNo': serial_no,
            'robotPwd': robotPwd

        }
        data = {}
        # 1为文件类型

        try:
            data = {}
            files = {'file': open(strs, 'rb')}
            r = requests.post(url=url, headers=header, files=files)
            ret = r.json().get('data')
            url = ret.get('url').replace('文件下载', nass_name)
            path = ret.get('path')
            logger.success(url, filename=path)
        except Exception as e:
            print(e)

    def main(self):
        name, time_year, time_month, flag = self.name_get()
        self.login()  # 登录
        main_iframe = self.page.wait_for_selector('//iframe[@id="frmMain"]').content_frame()  # 转换主要标签页
        main_iframe.wait_for_selector('//a[text()="新增"]', timeout=1000*60*2)
        if main_iframe.is_visible(f'//span[text()="{name}"]'):
            logger.info(f'"{name}"存在')
            if flag == 0:
                logger.info('项目已存在,若想进入校验流程，请在Excel的“A1”里加上“校验”')
                return
        else:
            logger.info(f'"{name}"不存在,可以用该名称创建新增项目')
            if flag == 1:
                logger.info('项目不存在，若想新建项目,请在Excel的“A1”里去掉“校验”')
                return
        print(name, time_year, time_month, flag)
        if flag == 0:
            self.new_create(name, time_year, time_month, main_iframe)  # 创建 新增项目
        file = self.file_download(name, main_iframe)  # 从"薪酬计算界面"进入 维护界面 里下载，留在导入与导出界面
        print(file)
        logger.info('进行Excel处理')
        rewrite = exceldeal(file)  # Excel 处理,只处理Excel，不对网页造成影响
        rewrite.main(name)  # Excel 处理,并删除 下载文件
        if flag == 0:
            self.update_excel(main_iframe, flag)  # 在 导入与导出界面 上传 已修改 的文件,点击全部计算，保存。回归"薪酬计算界面"
            time.sleep(5)
            file = self.file_download(name, main_iframe)  # 从"薪酬计算界面"进入 维护界面 里下载，留在导入与导出界面
            print(file)
            logger.info('进行第二次Excel处理')
            rewrite = exceldeal(file)  # Excel 处理
            rewrite.main(name)  # Excel 处理,并删除 下载文件
        self.page.close()  # 关闭浏览器
        self.context.close()  # 关闭浏览器
        self.updata_file(os.getcwd() + '\\data\\' + '导入模板.xlsx', nass_name='导入模板')
        self.send_email()  # 发送邮件


if __name__ == '__main__':
    # data_file = r'C:\Users\EDY\Documents\WeChat Files\wxid_oq5uzygli2oh22\FileStorage\File\2022-01\申雅线下表测试版(2)(1).xls'
    # deal = exceldeal(r'D:\python练习\data\薪酬数据导入_20220111 (2).xls')
    # deal.main(r'2021年8月冷链')

    setting = {
        # "recipientAddrs": '915709379@qq.com'
        "recipientAddrs":'qinyu@zhengming-sh.com'
    }

    logger.info('准备打开网页')
    with sync_playwright() as playwright:
        try:
            ZM = ZM(playwright)
            ZM.main()
        except Exception as e:
            logger.info(e)