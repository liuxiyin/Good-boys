import json
import os
import re
import time
import uuid
import pandas as pd
import requests
from Crypto.Hash import MD5
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from playwright.sync_api import sync_playwright
from loguru import logger
import datetime
import arrow
import shutil
import subprocess


class MA(object):

    def __init__(self):
        self.date_file = data_file
        self.result_file = os.getcwd() + '\\data\\' + str(uuid.uuid4()) + '.xlsx'
        self.air_url = 'https://gateway.uniner.com/track-j/api/airliftTrack/query'
        self.ship_url = 'https://gateway.uniner.com/track-j/api/shippingTrack/query'
        self.ups_url = 'https://gateway.uniner.com/track-j/api/expressTrack/query'

        try:
            self.server = subprocess.Popen(
                r"./ms-playwright/chromium-907428/chrome-win/chrome.exe --remote-debugging-port=9002")
        except:
            os.popen(
                './chrome --no-sandbox --disable-gpu --disable-dev-shm-usage --use-gl=desktop --window-size=1600,1024 --remote-debugging-port=9002')
        time.sleep(2)
        playwright = sync_playwright().start()
        self.browser = playwright.chromium.connect_over_cdp('http://localhost:9002')
        self.context = self.browser.contexts[0]
        self.page = self.context.pages[0]

        if os.path.exists(os.getcwd() + '\\data\\') is False:
            os.mkdir(os.getcwd() + '\\data\\')
        else:
            path_1 = os.getcwd() + '\\data\\'
            list_dir = os.listdir(path_1)
            for i in list_dir:
                file_name = os.path.join(path_1 + i)
                if os.path.isfile(file_name):
                    os.remove(file_name)
                elif os.path.isdir(file_name):
                    shutil.rmtree(file_name)

    def dhl_search(self, tracking_id):
        list2 = [None, None, None, None]
        try:
            self.page.goto('https://www.dhl.com/cn-zh/home.html')
            self.page.click('//h4/following-sibling::div[1]//input[@name="tracking-id"]')
            self.page.keyboard.press('Control+a')
            self.page.keyboard.press('Delete')
            self.page.type('//input[@name="tracking-id"]', str(tracking_id), delay=30)
            self.page.click('//button[@title="请填写此字段。"]')
            time.sleep(2)
        except:
            return list2
        for i in range(5):
            try:
                with self.page.expect_response(
                        f'https://www.dhl.com/utapi?trackingNumber={tracking_id}**',timeout=1000*30) as response_info:
                    self.page.click('//button[text()="追踪"]')
                    time.sleep(2)
                response = response_info.value
                response_dict = response.json()
                if 'shipments' in response_dict.keys():
                    events = response_dict.get('shipments')[0].get('events')
                    for j in events:
                        if '已取件' in j['description']:
                            list2[0] = j['timestamp'].replace("T",'')
                            list2[1] = j.get('location').get('address').get('addressLocality')
                        if '已派送' in j['description']:
                            list2[2] = j['timestamp'].replace("T",'')
                            list2[3] = j.get('location').get('address').get('addressLocality')
                    break
            except Exception as e:
                print(e)
                time.sleep(1)
        return list2

    def get_data_ups(self, num, list1):
        if str(num) == 'nan':
            return list1
        headers_post = {'Content-Type': 'application/json; charset=utf-8'}
        if '1Z' in str(num):
            data_post = {"firmType": 1, "waybillNo": num}  # ups
        elif str(num).isdigit() and len(str(num)) == 12:
            data_post = {"firmType": 2, "waybillNo": num}  # fedex
        # elif str(num).isdigit() and len(str(num)) == 10:
        #     data_post = {"firmType": 4, "waybillNo": num}  # DHL
        else:
            data_post = {"firmType": 3, "waybillNo": num}  # dpd
        for i in range(5):
            response = requests.post(self.ups_url, headers=headers_post, json=data_post)
            data_ups = response.json().get('data')
            status = response.json().get('status')
            if status == 500:
                time.sleep(0.95)
                if data_post['firmType'] == 3 and i == 2:
                    data_post['firmType'] = 4
                    status = '200'
                    break
                continue
            elif data_ups is None and status == 200:
                return list1
            elif data_ups.get('trackNodeVos') is None and status == 200:
                return list1
            else:
                break
        if status == 500:
            return list1
        if data_post['firmType'] == 1:  # ups
            for i in data_ups.get('trackNodeVos'):
                if i['nodeCode'] == 'OR' or '扫描' in i['sourceInfo']:
                    list1[13] = i['trackTime']
                    list1[14] = i['location']
                if 'Delivered' in i['sourceInfo']:
                    list1[15] = i['trackTime']
                    list1[16] = i['location']
        elif data_post['firmType'] == 2:  # fedex
            for i in data_ups.get('trackNodeVos'):
                if i['nodeCode'] == 'PU' or '已收取' in i['sourceInfo']:
                    list1[13] = i['trackTime']
                    list1[14] = i['location']
                if '"isDelivered":true' in i['sourceInfo']:
                    list1[15] = i['trackTime']
                    list1[16] = i['location']
        elif data_post['firmType'] == 3:  # dpd
            for i in data_ups.get('trackNodeVos'):
                str1 = i.get('sourceInfo')  # 获取存储信息的json
                str1 = self.str_deal(str1)
                if str(str1.get('eventCode')) == '009':
                    list1[13] = i['trackTime']
                    list1[14] = i['location']
                elif str(str1.get('eventCode')) == '001':
                    list1[15] = i['trackTime']
                    list1[16] = i['location']
        elif data_post['firmType'] == 4:  # DHL
            a = self.dhl_search(num)
            list1[13] = a[0]
            list1[14] = a[1]
            list1[15] = a[2]
            list1[16] = a[3]
        return list1

    def time_deal2(self, time_get):
        time_dict = {"January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6, "July": 7, "August": 8,
                     "September": 9, "October": 10, "November": 11, "December": 12}
        t = str(time_get).rsplit(" ")
        for month in time_dict:
            if t[2] in month:
                time1 = t[3] + '-' + str(time_dict[month]) + '-' + t[1] + ' ' + t[4]
                if int(t[4][:2]) < 12:
                    time1 = time1 + ' AM'
                else:
                    time1 = time1 + ' PM'
                return time1

    def time_deal(self, time_get):
        t = arrow.get(int(time_get)).to('local').shift(hours=-15)
        if t.hour < 12:
            t = t.format("YYYY-MM-DD HH:mm:ss") + " AM"
        else:
            t = t.format("YYYY-MM-DD HH:mm:ss") + " PM"
        return t

    def str_deal(self, s1):
        s2 = ' '
        s1 = s1.split()
        s1 = s2.join(s1)
        status = json.loads(s1)
        return status

    def get_data_air(self, num, list1):
        if str(num) == 'nan':
            return list1

        data_post = {"firmType": 1, "trackNumber": num}
        headers_post = {'Content-Type': 'application/json; charset=utf-8'}
        if num[:3] == '880':
            data_post["firmType"] = 1  # 海南航空
        elif num[:3] == '297':
            data_post["firmType"] = 2  # 台湾中华空运
        elif num[:3] == '180':
            data_post["firmType"] = 3  # 大韩航空
        elif num[:3] == '618':
            data_post["firmType"] = 4  # 新加坡航空
        elif num[:3] == '176':
            data_post["firmType"] = 5  # 阿联酋航空
        elif num[:3] == '695':
            data_post["firmType"] = 6  # 台湾长荣航空
        elif num[:3] == '157':
            data_post["firmType"] = 7  # 卡塔尔航空
        elif num[:3] == '988':
            data_post["firmType"] = 8  # 韩亚航空
        elif num[:3] == '160':
            data_post["firmType"] = 9  # 韩亚航空

        for i in range(3):
            response = requests.post(self.air_url, headers=headers_post, json=data_post)
            data_air = response.json().get("data")
            status = response.json().get("status")
            if status == 500:
                time.sleep(1)
                continue
            elif data_air is None and status == 200:  # 如果读取内容为空
                return list1
            elif data_air.get('trackNodeVos') is None and status == 200:
                return list1
            else:
                break

        if status == 500:
            return list1
        if data_post["firmType"] == 1:  # 海南航空
            for i in data_air.get("trackNodeVos"):
                if i['nodeCode'] == 'DLV' and list1[11] is None:
                    list1[11] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d')
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD' and list1[9] is None:
                    list1[9] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d')
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS' and list1[7] is None:
                    list1[7] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d')
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR' and list1[5] is None:
                    list1[5] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP':
                    list1[3] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d')
                    list1[4] = i['location']
        elif data_post["firmType"] == 2:  # 台湾中华空运
            year = datetime.datetime.now().year
            for i in data_air.get("trackNodeVos"):
                if i['nodeCode'] == 'DLV' and list1[11] is None:
                    list1[11] = datetime.datetime.strptime(i['trackTime'].split()[0], '%d%b').strftime(f'{year}%m%d')
                    if int(list1[11])>int(datetime.datetime.now().strftime('%Y%m%d')):
                        list1[11] = datetime.datetime.strptime(list1[11], '%Y%m%d').strftime(f'{year - 1}/%m/%d')
                    else:
                        list1[11] = datetime.datetime.strptime(list1[11], '%Y%m%d').strftime(f'{year}/%m/%d')
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD' and list1[9] is None:
                    list1[9] = datetime.datetime.strptime(i['trackTime'].split()[0], '%d%b').strftime(f'{year}%m%d')
                    if int(list1[9]) > int(datetime.datetime.now().strftime('%Y%m%d')):
                        list1[9] = datetime.datetime.strptime(list1[9], '%Y%m%d').strftime(f'{year - 1}/%m/%d')
                    else:
                        list1[9] = datetime.datetime.strptime(list1[9], '%Y%m%d').strftime(f'{year}/%m/%d')
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS' and list1[7] is None:
                    list1[7] = datetime.datetime.strptime(i['trackTime'].split()[0], '%d%b').strftime(f'{year}%m%d')
                    if int(list1[7]) > int(datetime.datetime.now().strftime('%Y%m%d')):
                        list1[7] = datetime.datetime.strptime(list1[7], '%Y%m%d').strftime(f'{year - 1}/%m/%d')
                    else:
                        list1[7] = datetime.datetime.strptime(list1[7], '%Y%m%d').strftime(f'{year}/%m/%d')
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR' and list1[5] is None:
                    list1[5] = datetime.datetime.strptime(i['trackTime'].split()[0], '%d%b').strftime(f'{year}%m%d')
                    if int(list1[5]) > int(datetime.datetime.now().strftime('%Y%m%d')):
                        list1[5] = datetime.datetime.strptime(list1[5], '%Y%m%d').strftime(f'{year - 1}/%m/%d')
                    else:
                        list1[5] = datetime.datetime.strptime(list1[5], '%Y%m%d').strftime(f'{year}/%m/%d')
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP':
                    list1[3] = datetime.datetime.strptime(i['trackTime'].split()[0], '%d%b').strftime(f'{year}%m%d')
                    if int(list1[3]) > int(datetime.datetime.now().strftime('%Y%m%d')):
                        list1[3] = datetime.datetime.strptime(list1[3], '%Y%m%d').strftime(f'{year - 1}/%m/%d')
                    else:
                        list1[3] = datetime.datetime.strptime(list1[3], '%Y%m%d').strftime(f'{year}/%m/%d')
                    list1[4] = i['location']
        elif data_post["firmType"] == 3:  # 大韩航空
            for i in data_air.get("trackNodeVos"):
                if i['nodeCode'] == 'DLV' and list1[11] is None:
                    list1[11] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD' and list1[9] is None:
                    list1[9] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS' and list1[7] is None:
                    list1[7] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR' and list1[5] is None:
                    list1[5] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP':
                    list1[3] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[4] = i['location']
        elif data_post["firmType"] == 4:  # 新加坡航空
            for i in data_air.get("trackNodeVos"):
                if i['nodeCode'] == 'DLV':
                    list1[11] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y').strftime('%Y/%m/%d')
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD':
                    list1[9] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y').strftime('%Y/%m/%d')
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS':
                    list1[7] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y').strftime('%Y/%m/%d')
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR':
                    list1[5] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP' and list1[3] is None:
                    list1[3] = datetime.datetime.strptime(i['trackTime'], '%d %b %Y').strftime('%Y/%m/%d')
                    list1[4] = i['location']
        elif data_post["firmType"] == 5:  # 阿联酋航空
            for i in data_air.get("trackNodeVos"):
                time1 = i['trackTime'].split(',')[0].replace("]", "").replace("[", "").replace('"', '')
                if i['nodeCode'] == 'DLV' and list1[11] is None:
                    list1[11] = datetime.datetime.strptime(time1, '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD' and list1[9] is None:
                    list1[9] = datetime.datetime.strptime(time1, '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS' and list1[7] is None:
                    list1[7] = datetime.datetime.strptime(time1, '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR' and list1[5] is None:
                    list1[5] = datetime.datetime.strptime(time1, '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP':
                    list1[3] = datetime.datetime.strptime(time1, '%d %b %Y %H:%M').strftime('%Y/%m/%d')
                    list1[4] = i['location']
        elif data_post["firmType"] == 6:  # 台湾长荣航空
            for i in data_air.get("trackNodeVos"):
                if i['nodeCode'] == 'DLV' and list1[11] is None:
                    list1[11] = datetime.datetime.strptime(i['trackTime'], '%Y/%m/%d %H:%M').strftime('%Y/%m/%d')
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD' and list1[9] is None:
                    list1[9] = datetime.datetime.strptime(i['trackTime'], '%Y/%m/%d %H:%M').strftime('%Y/%m/%d')
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS' and list1[7] is None:
                    list1[7] = datetime.datetime.strptime(i['trackTime'], '%Y/%m/%d %H:%M').strftime('%Y/%m/%d')
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR' and list1[5] is None:
                    list1[5] = datetime.datetime.strptime(i['trackTime'], '%Y/%m/%d %H:%M').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP':
                    list1[3] = datetime.datetime.strptime(i['trackTime'], '%Y/%m/%d %H:%M').strftime('%Y/%m/%d')
                    list1[4] = i['location']
        elif data_post["firmType"] == 7:  # 卡塔尔航空
            for i in data_air.get("trackNodeVos"):
                if i['nodeCode'] == 'DLV' and list1[11] is None:
                    list1[11] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y').strftime('%Y/%m/%d')
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD' and list1[9] is None:
                    list1[9] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y').strftime('%Y/%m/%d')
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS' and list1[7] is None:
                    list1[7] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y').strftime('%Y/%m/%d')
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR' and list1[5] is None:
                    list1[5] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP':
                    list1[3] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y').strftime('%Y/%m/%d')
                    list1[4] = i['location']
        elif data_post["firmType"] == 8:  # 韩亚航空
            for i in data_air.get("trackNodeVos"):
                if i['nodeCode'] == 'DLV':
                    list1[11] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M').strftime('%Y/%m/%d')
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD':
                    list1[9] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M').strftime('%Y/%m/%d')
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS':
                    list1[7] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M').strftime('%Y/%m/%d')
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR':
                    list1[5] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP' and list1[3] is None:
                    list1[3] = datetime.datetime.strptime(i['trackTime'], '%Y-%m-%d %H:%M').strftime('%Y/%m/%d')
                    list1[4] = i['location']
        elif data_post["firmType"] == 9:  # 韩亚航空
            for i in data_air.get("trackNodeVos"):
                if i['nodeCode'] == 'DLV':
                    list1[11] = i['trackTime']
                    list1[12] = i['location']
                elif i['nodeCode'] == 'NFD':
                    list1[9] = i['trackTime']
                    list1[10] = i['location']
                elif i['nodeCode'] == 'CUS':
                    list1[7] = i['trackTime']
                    list1[8] = i['location']
                elif i['nodeCode'] == 'ARR':
                    list1[5] = i['trackTime']
                    list1[6] = i['location']
                elif i['nodeCode'] == 'DEP' and list1[3] is None:
                    list1[3] = i['trackTime']
                    list1[4] = i['location']
        return list1

    def get_data_maston_ship(self,num,list1):
        if str(num) == 'nan':
            return list1

        data_post = {"firmType": 1, "containerNumber": '5iqv8q', 'ladingNumber': num}
        headers_post = {'Content-Type': 'application/json; charset=utf-8'}
        if num[:4] == 'MATS':  # 美森
            data_post["firmType"] = 1
        elif num[:2] == 'CN' or num[:4] == 'EGLV':  # 达飞
            data_post["firmType"] = 2
        elif num[:4] == 'TEMU' or num[:4] == 'ZCSU':  # 以星
            data_post["firmType"] = 3
        elif num[:4] == 'COSU':  # cosco
            data_post["firmType"] = 4
        elif num[:4] == 'ONEY':  # one
            data_post["firmType"] = 5
        elif str(num).isdigit() is True:  # ct EMC
            data_post["firmType"] = 6
        elif num[:4] == 'OOLU':  # oocl
            data_post["firmType"] = 7
        elif num[:4] == 'JJCO':  # jj
            data_post["firmType"] = 8
        for i in range(3):
            response = requests.post(self.ship_url, headers=headers_post, json=data_post)
            data_ship = response.json().get("data")
            status = response.json().get("status")
            if status == 500:
                time.sleep(0.95)
                continue
            elif data_ship is None and status == 200:  # 如果读取内容为空
                return list1
            elif data_ship.get('trackNodeVos') is None and status == 200:
                return list1
            else:
                break

        if status == 500:
            return list1
        if data_post["firmType"] == 1:  # 美森
            for i in range(len(data_ship.get("trackNodeVos"))):
                time1 = data_ship.get("trackNodeVos")[i].get('trackTime')  # 获取时间
                time1 = self.time_deal(time1)
                str1 = data_ship.get("trackNodeVos")[i].get('sourceInfo')  # 获取存储信息的json
                str1 = self.str_deal(str1)
                status = str1.get('status')
                if 'OUTGATE' in status:
                    if 'OUTGATE to SHIPPERS TRANSPORT EXPRESS' in status: continue
                    list1[11] = time1
                    list1[12] = status
                elif 'AVAILABLE' in status:
                    list1[9] = time1
                    list1[10] = status
                elif 'DISCHARGE' in status:
                    list1[7] = time1
                    list1[8] = status
                elif 'HOLD' in status and 'Customs' in status:
                    list1[5] = time1
                    list1[6] = status
                elif 'LOAD' in status:
                    list1[3] = time1
                    list1[4] = status
        elif data_post["firmType"] == 2:  # 达飞
            for i in data_ship.get("trackNodeVos"):
                if 'Loaded on board' in i['sourceInfo']:
                    list1[3] = self.time_deal2(i['trackTime'])
                    list1[4] = str(i['location']).replace('</p>', '').replace('<p>', '')
                elif 'Discharged' in i['sourceInfo']:
                    list1[5] = self.time_deal2(i['trackTime'])
                    list1[6] = str(i['location']).replace('</p>', '').replace('<p>', '')
                elif 'Container' in i['sourceInfo']:
                    list1[11] = self.time_deal2(i['trackTime'])
                    list1[12] = str(i['location']).replace('</p>', '').replace('<p>', '')
        elif data_post["firmType"] == 3:  # 以星
            for i in data_ship.get("trackNodeVos"):
                if 'Container was loaded' in i['sourceInfo']:
                    list1[3] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y %H:%M').strftime('%Y/%m/%d')
                    list1[4] = i['location']
                elif 'Vessel arrival' in i['sourceInfo']:
                    list1[5] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y %H:%M').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif 'Customs' in i['sourceInfo']:
                    list1[7] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y %H:%M').strftime('%Y/%m/%d')
                    list1[8] = i['location']
                elif 'available' in i['sourceInfo']:
                    list1[9] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y %H:%M').strftime('%Y/%m/%d')
                    list1[10] = i['location']
                elif 'Import' in i['sourceInfo']:
                    list1[11] = datetime.datetime.strptime(i['trackTime'], '%d-%b-%Y %H:%M').strftime('%Y/%m/%d')
                    list1[12] = i['location']
        elif data_post["firmType"] == 4:  # cosco
            for i in data_ship.get("trackNodeVos"):
                if '离开' in i['nodeName']:
                    list1[3] = i['trackTime']
                    list1[4] = i['location']
                elif '到达' in i['nodeName']:
                    list1[5] = i['trackTime']
                    list1[6] = i['location']
                elif '提柜' in i['nodeName']:
                    list1[11] = i['trackTime']
                    list1[12] = i['location']
        elif data_post["firmType"] == 5:  # one
            for i in data_ship.get("trackNodeVos"):
                if '离开' in i['nodeName']:
                    list1[3] = i['trackTime']
                    list1[4] = i['location']
                elif '到达' in i['nodeName']:
                    list1[5] = i['trackTime']
                    list1[6] = i['location']
                elif '提柜' in i['nodeName']:
                    list1[11] = i['trackTime']
                    list1[12] = i['location']
        elif data_post["firmType"] == 6:  # ct EMC
            for i in data_ship.get("trackNodeVos"):
                if '装船' in i['nodeName']:
                    list1[3] = datetime.datetime.strptime(i['trackTime'], '%b-%d-%Y').strftime('%Y/%m/%d')
                    list1[4] = i['location']
                elif '卸船' in i['nodeName']:
                    list1[5] = datetime.datetime.strptime(i['trackTime'], '%b-%d-%Y').strftime('%Y/%m/%d')
                    list1[6] = i['location']
                elif '提柜' in i['nodeName']:
                    list1[11] = datetime.datetime.strptime(i['trackTime'], '%b-%d-%Y').strftime('%Y/%m/%d')
                    list1[12] = i['location']
        elif data_post["firmType"] == 7:  # OOCL
            for i in data_ship.get("trackNodeVos"):
                time1 = i['trackTime'].split(',')[0]
                time1 = datetime.datetime.strptime(time1, '%d %b %Y').strftime('%Y/%m/%d')
                if '离开' in i['nodeName']:
                    list1[3] = time1
                    list1[4] = i['location']
                elif 'Arrived' in i['sourceInfo']:
                    list1[5] = time1
                    list1[6] = i['location']
                elif 'Available' in i['sourceInfo']:
                    list1[9] = time1
                    list1[10] = i['location']
                elif 'Delivery' in i['sourceInfo']:
                    list1[11] = time1
                    list1[12] = i['location']
        elif data_post["firmType"] == 8:  # ct EMC
            for i in data_ship.get("trackNodeVos"):
                if '装船' in i['nodeName']:
                    list1[3] = i['trackTime']
                    list1[4] = i['location']
                elif '卸船' in i['nodeName']:
                    list1[5] = i['trackTime']
                    list1[6] = i['location']
                elif '提柜' in i['nodeName']:
                    list1[11] = i['trackTime']
                    list1[12] = i['location']
        return list1

    def get_data(self, file):
        data_read = pd.read_excel(file, sheet_name='Sheet1', dtype=str)
        data_read = data_read.loc[:, ['DP单号', '提单号', '快递单号']]
        data_save = {'DP单号': [], '提单号': [], '快递单号': [], '起飞': [], 'Unnamed: 4': [], '抵达': [], 'Unnamed: 6': [],
                     '清关': [], 'Unnamed: 8': [], '提货通知': [], 'Unnamed: 10': [], '提货': [], 'Unnamed: 12': [], '快递提取': [],
                     'Unnamed: 14': [], '快递签收': [], 'Unnamed: 16': [], '起飞/开船-抵达/到港时效（自然日）': [],
                     '抵达/到港-目的国海关放行（工作日）': [], '目的国海关放行-快递提取（工作日）': [], '快递提取-快递签收（工作日）': []}
        logger.info('表格读取完毕')
        for index, value in data_read.iterrows():
            a = [value['DP单号'], value['提单号'], value['快递单号'], None, None, None, None, None, None, None, None, None, None,
                 None, None, None, None, None, None, None, None]  # 提单号-未查询到（0-14）
            if str(value['DP单号']) != 'nan':
                if len(data_save['DP单号']) > 1 and data_save['DP单号'][-1] == value['DP单号']:
                    for i, j in zip(range(13), data_save):
                        if i > 2: a[i] = data_save[j][-1]
                else:
                    for i in range(5):
                        try:
                            a = self.get_data_air(value['DP单号'], a)
                            break
                        except Exception as e:
                            logger.info(e)
            else:
                if len(data_save['提单号']) > 1 and data_save['提单号'][-1] == value['提单号']:
                    for i, j in zip(range(13), data_save):
                        if i > 2: a[i] = data_save[j][-1]
                else:
                    for i in range(5):
                        try:
                            a = self.get_data_maston_ship(value['提单号'], a)
                            break
                        except Exception as e:
                            logger.info(e)
            for i in range(5):
                try:
                    a = self.get_data_ups(value['快递单号'], a)
                    break
                except Exception as e:
                    logger.info(e)
            if a[3] is not None and a[5] is not None:
                a[17] = a[3] + ' —— ' + a[5]
            if a[7] is not None and a[5] is not None:
                a[18] = a[5] + ' —— ' + a[7]
            elif a[9] is not None and a[5] is not None:
                a[18] = a[5] + ' —— ' + a[9]
            if a[13] is not None and a[7] is not None:
                a[19] = a[7] + ' —— ' + a[13]
            elif a[13] is not None and a[9] is not None:
                a[19] = a[9] + ' —— ' + a[13]
            if a[13] is not None and a[15] is not None:
                a[20] = a[13] + ' —— ' + a[15]

            logger.info(f'{index},{a}')
            for i, j in zip(a, data_save):
                data_save[j].append(i)
        self.page.close()
        self.browser.close()
        # self.server.kill()
        data_save = pd.DataFrame(data=data_save)
        data_save.to_excel(self.result_file, index=False)

    def change_excel(self, filename):
        wb = load_workbook(filename)
        ws = wb.active
        fill = PatternFill(patternType="solid", start_color="00b0f0")
        list1 = ['A1:A2', 'B1:B2', 'C1:C2', 'D1:E1', 'F1:G1', 'H1:I1', 'J1:K1', 'L1:M1', 'N1:O1', 'P1:Q1', 'R1:R2',
                 'S1:S2', 'T1:T2', 'U1:U2']
        for i in list1:
            ws.merge_cells(i)  # 合并单元格
        list2 = [['A1', 'DP单号'], ['B1', '提单号'], ['C1', '快递单号'], ['D1', '起飞'], ['D2', '日期'], ['E2', '描述'], ['F1', '抵达'],
                 ['F2', '日期'], ['G2', '描述'], ['H1', '清关'], ['H2', '日期'], ['I2', '描述'], ['J1', '提货通知'], ['J2', '日期'],
                 ['K2', '描述'], ['L1', '提货'], ['L2', '日期'], ['M2', '描述'], ['N1', '快递提取'], ['N2', '日期'], ['O2', '描述'],
                 ['P1', '快递签收'], ['P2', '日期'], ['Q2', '描述'], ['R1', '起飞/开船-抵达/到港时效（自然日）'],
                 ['S1', '抵达/到港-目的国海关放行（工作日）'], ['T1', '目的国海关放行-快递提取（工作日）'], ['U1', '快递提取-快递签收（工作日）']]
        for i in list2:
            ws[i[0]] = i[1]  # 写入表头
        list3 = ['A1', 'B1', 'C1', 'D1', 'D2', 'E2', 'F1', 'F2', 'G2', 'H1', 'H2', 'I2', 'J1', 'J2', 'K2', 'L1', 'L2',
                 'M2', 'N1', 'N2', 'O2', 'P1', 'P2', 'Q2', 'R1', 'S1', 'T1', 'U1']
        for i in list3:
            alignment = Alignment(horizontal='center', vertical='center')
            ws[i].alignment = alignment
        list4 = ['A1', 'B1', 'C1', 'N1', 'N2', 'O2', 'P1', 'P2', 'Q2']
        for i in list4:
            ws[i].fill = fill  # 表头颜色替换
        wb.save(filename)
        wb.close()

    def updata_file(self, strs, nass_name='文件'):
        from loguru import logger
        # robotNo = serial_no
        # robotPwd = robot_secret

        obj = MD5.new()
        obj.update(robot_secret.encode("utf-8"))  # gb2312 Or utf-8
        robotPwd = obj.hexdigest()

        url = f'{base_url}/worker/internal/file/upload/result'
        # url = f'{base_url}/file/upload/{robotNo}/{robotPwd}'
        header = {
            'robotNo': serial_no,
            'robotPwd': robotPwd

        }
        data = {}
        # 1为文件类型

        try:
            data = {}
            files = {'file': open(strs, 'rb')}
            r = requests.post(url=url, headers=header, files=files)
            ret = r.json().get('data')
            url = ret.get('url').replace('文件下载', nass_name)
            path = ret.get('path')
            logger.success(url, filename=path)
        except Exception as e:
            print(e)

    def main(self):
        self.get_data(data_file)
        self.change_excel(self.result_file)
        self.updata_file(self.result_file)


if __name__ == '__main__':
    # data_file = r'./空运模板表.xlsx'
    MA = MA()
    MA.main()
